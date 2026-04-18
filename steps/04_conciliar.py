"""
steps/04_conciliar.py — Conciliación Payway × Zetti
Doblen Solutions x Farmacias del Pueblo

Lógica idéntica al 03_conciliar.py original.
Diferencia: recibe rutas de archivos por parámetro en lugar de buscarlos
en carpetas locales — así encaja en el pipeline de Railway.
"""

import os
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Mapa Zetti → Payway (idéntico al original)
# ---------------------------------------------------------------------------
MAPA_SUCURSALES = {
    "FCIA ALBERDI NQN":                  "alberdi",
    "FCIA ALLEN RN":                     "allen",
    "FCIA ALTO COMAHUE NQN":             "alto comahue",
    "FCIA BELGRANO Y BOQUET ROLDAN NQN": "belgrano",
    "FCIA CENTENARIO NQN":               "centenario",
    "FCIA CENTRAL NQN":                  "central",
    "FCIA CIPOLLETTI RN":                "cipolletti",
    "FCIA COMPLEJO OESTE NQN":           "oeste",
    "FCIA ITALIA NQN":                   "italia",
    "FCIA JUAN B JUSTO NQN":             "j b justo",
    "FCIA JUMBO NQN":                    "jumbo",
    "FCIA LA ANONIMA NQN":               "anonima",
    "FCIA MITRE NQN":                    "mitre",
    "FCIA PERITO MORENO NQN":            "perito moreno",
    "FCIA PERTICONE NQN":                "perticone",
    "FCIA PLOTTIER NQN":                 "plottier",
    "FCIA PLOTTIER 2 NQN":               "plottier",
    "FCIA PLOTTIER 3 NQN":               "plottier",
    "FCIA SAN MARTIN NQN":               "san martin",
    "FCIA SENILLOSA":                    "senillosa",
    "FCIA ZAPALA NQN":                   "zapala",
}


# ---------------------------------------------------------------------------
# Normalización de tarjetas (idéntica al original)
# ---------------------------------------------------------------------------
def normalizar_tarjeta(nombre: str) -> str:
    if pd.isna(nombre):
        return "DESCONOCIDA"
    n = str(nombre).upper().strip()

    if "VISA" in n and any(x in n for x in ("DEBITO", "DÉBITO", "ELECTRON")):
        return "VISA ELECTRON"
    if "VISA" in n and "PREPAGO" in n:
        return "VISA ELECTRON"
    if "VISA" in n:
        return "VISA"

    if ("MASTERCARD" in n or "MASTER CARD" in n) and any(x in n for x in ("DEBIT", "DÉBITO", "DEBITO")):
        return "MASTERCARD DEBITO"
    if ("MASTERCARD" in n or "MASTER CARD" in n) and "PREPAGO" in n:
        return "MASTERCARD DEBITO"
    if "MASTERCARD" in n or "MASTER CARD" in n:
        return "MASTERCARD"

    if "CABAL" in n and any(x in n for x in ("DEBITO", "DÉBITO", "24")):
        return "CABAL 24"
    if "CABAL" in n:
        return "CABAL"

    if "NARANJA" in n:
        return "NARANJA CREDITO"
    if "AMERICAN EXPRESS" in n or "AMEX" in n:
        return "AMERICAN EXPRESS"
    if "CREDIGUIA" in n:
        return "CREDIGUIA"
    if "CONFIABLE" in n:
        return "CONFIABLE CREDITO"
    if "MAESTRO" in n:
        return "MAESTRO"
    if "MERCADO PAGO" in n:
        return "TARJETA MERCADO PAGO QR"
    if "MODO" in n:
        return "TARJETA MODO"

    return n


# ---------------------------------------------------------------------------
# Estilos (idénticos al original)
# ---------------------------------------------------------------------------
_thin  = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

COLOR_HEADER  = "1F4E79"
COLOR_OK      = "E2EFDA"
COLOR_WARN    = "FFF2CC"
COLOR_ERROR   = "FCE4D6"
COLOR_MISSING = "F2F2F2"
COLOR_GRIS    = "F9F9F9"
COLOR_TOTAL   = "BDD7EE"

def _hdr(cell, texto=None):
    if texto is not None:
        cell.value = texto
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def _dat(cell, bg=None, bold=False, align="right"):
    cell.font      = Font(name="Arial", bold=bold, size=9)
    cell.fill      = PatternFill("solid", start_color=bg or "FFFFFF")
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")

def _num(cell, bg=None, bold=False):
    _dat(cell, bg=bg, bold=bold, align="right")
    cell.number_format = "#,##0.00"

def _pct(cell, bg=None):
    _dat(cell, bg=bg, align="right")
    cell.number_format = "0.0%"

def color_dif(pct_val):
    if pct_val is None:
        return COLOR_MISSING
    ap = abs(pct_val)
    if ap < 0.01:   return COLOR_OK
    elif ap < 0.05: return COLOR_WARN
    else:           return COLOR_ERROR


# ---------------------------------------------------------------------------
# Lectura de datos (idéntica al original)
# ---------------------------------------------------------------------------
def leer_payway(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Carga de lotes", header=1, dtype=str, engine="openpyxl")
    df.columns = df.columns.str.strip()

    rename = {
        "Fecha":         "fecha",
        "Sucursal":      "sucursal",
        "Tarjeta":       "tarjeta_raw",
        "Transacciones": "txn_payway",
        "Importe Total": "monto_payway",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    df = df[df["fecha"].notna() & ~df["fecha"].str.lower().str.contains("subtotal|total", na=False)]

    df["monto_payway"] = pd.to_numeric(df["monto_payway"], errors="coerce").fillna(0)
    df["txn_payway"]   = pd.to_numeric(df["txn_payway"],   errors="coerce").fillna(0)
    df["tarjeta"]      = df["tarjeta_raw"].apply(normalizar_tarjeta)
    df["sucursal"]     = df["sucursal"].str.strip().str.lower()

    return df.groupby(["sucursal", "tarjeta"]).agg(
        txn_payway=("txn_payway",   "sum"),
        monto_payway=("monto_payway", "sum"),
    ).reset_index()


def leer_zetti(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";", dtype=str, encoding="utf-8-sig")
    df.columns = df.columns.str.strip()

    rename = {
        "sucursal_nombre": "sucursal_zetti",
        "tarjeta_nombre":  "tarjeta_raw",
        "cupones":         "txn_zetti",
        "monto_total":     "monto_zetti",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})

    df["monto_zetti"] = pd.to_numeric(df["monto_zetti"], errors="coerce").fillna(0)
    df["txn_zetti"]   = pd.to_numeric(df["txn_zetti"],   errors="coerce").fillna(0)
    df["tarjeta"]     = df["tarjeta_raw"].apply(normalizar_tarjeta)
    df["sucursal"]    = df["sucursal_zetti"].map(
        lambda x: MAPA_SUCURSALES.get(str(x).strip(), str(x).strip().lower())
    )

    return df.groupby(["sucursal", "tarjeta"]).agg(
        txn_zetti=("txn_zetti",     "sum"),
        monto_zetti=("monto_zetti", "sum"),
        sucursal_zetti=("sucursal_zetti", "first"),
    ).reset_index()


# ---------------------------------------------------------------------------
# Conciliación (idéntica al original)
# ---------------------------------------------------------------------------
def conciliar(df_pay: pd.DataFrame, df_zet: pd.DataFrame) -> pd.DataFrame:
    merged = pd.merge(df_pay, df_zet, on=["sucursal", "tarjeta"], how="outer")
    merged["monto_payway"] = merged["monto_payway"].fillna(0)
    merged["monto_zetti"]  = merged["monto_zetti"].fillna(0)
    merged["txn_payway"]   = merged["txn_payway"].fillna(0)
    merged["txn_zetti"]    = merged["txn_zetti"].fillna(0)
    merged["diferencia"]   = merged["monto_payway"] - merged["monto_zetti"]

    def pct_dif(row):
        base = max(abs(row["monto_payway"]), abs(row["monto_zetti"]))
        return None if base == 0 else row["diferencia"] / base

    merged["pct_diferencia"] = merged.apply(pct_dif, axis=1)

    def estado(row):
        if row["monto_payway"] == 0: return "Solo Zetti"
        if row["monto_zetti"]  == 0: return "Solo Payway"
        ap = abs(row["pct_diferencia"] or 0)
        if ap < 0.01:   return "OK"
        elif ap < 0.05: return "Dif. menor"
        else:           return "Revisar"

    merged["estado"] = merged.apply(estado, axis=1)
    return merged.sort_values(["sucursal", "tarjeta"]).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel de conciliación (idéntico al original)
# ---------------------------------------------------------------------------
def escribir_excel(df: pd.DataFrame, output_path: str, fecha: date):
    wb = Workbook()

    # Hoja 1 — Conciliación
    ws = wb.active
    ws.title        = "Conciliación"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value     = f"Conciliación Payway × Zetti — {fecha.strftime('%d/%m/%Y')}"
    t.font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color=COLOR_HEADER)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = [
        "Sucursal", "Tarjeta (normalizada)",
        "Txn Payway", "Monto Payway ($)",
        "Txn Zetti",  "Monto Zetti ($)",
        "Diferencia ($)", "Dif. %", "Estado",
        "Sucursal Zetti (original)",
    ]
    for ci, h in enumerate(headers, 1):
        _hdr(ws.cell(row=2, column=ci), h)
    ws.row_dimensions[2].height = 28

    for ri, row in enumerate(df.itertuples(index=False), 3):
        bg       = color_dif(row.pct_diferencia)
        sucursal = str(row.sucursal).title() if pd.notna(row.sucursal) else ""
        tarjeta  = str(row.tarjeta)          if pd.notna(row.tarjeta)  else ""
        suc_orig = str(row.sucursal_zetti)   if pd.notna(row.sucursal_zetti) else ""

        ws.cell(row=ri, column=1,  value=sucursal);          _dat(ws.cell(row=ri,column=1),  bg=bg, align="left")
        ws.cell(row=ri, column=2,  value=tarjeta);           _dat(ws.cell(row=ri,column=2),  bg=bg, align="left")
        ws.cell(row=ri, column=3,  value=row.txn_payway);    _num(ws.cell(row=ri,column=3),  bg=bg); ws.cell(row=ri,column=3).number_format="#,##0"
        ws.cell(row=ri, column=4,  value=row.monto_payway);  _num(ws.cell(row=ri,column=4),  bg=bg)
        ws.cell(row=ri, column=5,  value=row.txn_zetti);     _num(ws.cell(row=ri,column=5),  bg=bg); ws.cell(row=ri,column=5).number_format="#,##0"
        ws.cell(row=ri, column=6,  value=row.monto_zetti);   _num(ws.cell(row=ri,column=6),  bg=bg)
        ws.cell(row=ri, column=7,  value=row.diferencia);    _num(ws.cell(row=ri,column=7),  bg=bg, bold=abs(row.diferencia) > 1000)
        ws.cell(row=ri, column=8,  value=row.pct_diferencia); _pct(ws.cell(row=ri,column=8), bg=bg)
        ws.cell(row=ri, column=9,  value=row.estado);         _dat(ws.cell(row=ri,column=9), bg=bg, bold=True, align="center")
        ws.cell(row=ri, column=10, value=suc_orig);           _dat(ws.cell(row=ri,column=10),bg=COLOR_GRIS, align="left")

    last    = 2 + len(df)
    tot_row = last + 1
    ws.merge_cells(f"A{tot_row}:B{tot_row}")
    tl = ws.cell(row=tot_row, column=1, value="TOTAL")
    tl.font = Font(name="Arial", bold=True, size=10)
    tl.fill = PatternFill("solid", start_color=COLOR_TOTAL)
    tl.border = BORDER; tl.alignment = Alignment(horizontal="left")

    for ci, _ in [(3,"txn_payway"),(4,"monto_payway"),(5,"txn_zetti"),(6,"monto_zetti"),(7,"diferencia")]:
        cell = ws.cell(row=tot_row, column=ci, value=f"=SUM({get_column_letter(ci)}3:{get_column_letter(ci)}{last})")
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.fill = PatternFill("solid", start_color=COLOR_TOTAL)
        cell.border = BORDER; cell.alignment = Alignment(horizontal="right")
        cell.number_format = "#,##0.00" if ci not in (3, 5) else "#,##0"

    for ci in [8, 9, 10]:
        ws.cell(row=tot_row, column=ci).fill   = PatternFill("solid", start_color=COLOR_TOTAL)
        ws.cell(row=tot_row, column=ci).border = BORDER

    for col, w in {"A":16,"B":20,"C":12,"D":18,"E":12,"F":18,"G":18,"H":10,"I":12,"J":26}.items():
        ws.column_dimensions[col].width = w

    # Hoja 2 — Para revisar
    ws2 = wb.create_sheet("Para revisar")
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:I1")
    t2 = ws2["A1"]
    t2.value     = "Items para revisar — diferencia > 1% o solo en un sistema"
    t2.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t2.fill      = PatternFill("solid", start_color="C55A11")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    for ci, h in enumerate(headers[:9], 1):
        _hdr(ws2.cell(row=2, column=ci), h)
    ws2.row_dimensions[2].height = 26

    df_revisar = df[df["estado"].isin(["Revisar", "Dif. menor", "Solo Payway", "Solo Zetti"])]
    for ri, row in enumerate(df_revisar.itertuples(index=False), 3):
        bg = color_dif(row.pct_diferencia)
        ws2.cell(row=ri, column=1, value=str(row.sucursal).title()); _dat(ws2.cell(row=ri,column=1), bg=bg, align="left")
        ws2.cell(row=ri, column=2, value=str(row.tarjeta));          _dat(ws2.cell(row=ri,column=2), bg=bg, align="left")
        ws2.cell(row=ri, column=3, value=row.txn_payway);   _num(ws2.cell(row=ri,column=3), bg=bg); ws2.cell(row=ri,column=3).number_format="#,##0"
        ws2.cell(row=ri, column=4, value=row.monto_payway); _num(ws2.cell(row=ri,column=4), bg=bg)
        ws2.cell(row=ri, column=5, value=row.txn_zetti);    _num(ws2.cell(row=ri,column=5), bg=bg); ws2.cell(row=ri,column=5).number_format="#,##0"
        ws2.cell(row=ri, column=6, value=row.monto_zetti);  _num(ws2.cell(row=ri,column=6), bg=bg)
        ws2.cell(row=ri, column=7, value=row.diferencia);   _num(ws2.cell(row=ri,column=7), bg=bg, bold=True)
        ws2.cell(row=ri, column=8, value=row.pct_diferencia); _pct(ws2.cell(row=ri,column=8), bg=bg)
        ws2.cell(row=ri, column=9, value=row.estado);        _dat(ws2.cell(row=ri,column=9), bg=bg, bold=True, align="center")

    for col, w in {"A":16,"B":20,"C":12,"D":18,"E":12,"F":18,"G":18,"H":10,"I":12}.items():
        ws2.column_dimensions[col].width = w

    # Hoja 3 — Leyenda
    ws3 = wb.create_sheet("Leyenda")
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 50

    ws3.merge_cells("A1:B1")
    tl3 = ws3["A1"]
    tl3.value = "Leyenda de colores — Conciliación"
    tl3.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    tl3.fill  = PatternFill("solid", start_color=COLOR_HEADER)
    tl3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24

    leyenda = [
        (COLOR_OK,      "OK",          "Diferencia < 1% — concilia correctamente"),
        (COLOR_WARN,    "Dif. menor",  "Diferencia entre 1% y 5% — revisar si es QR/prepago"),
        (COLOR_ERROR,   "Revisar",     "Diferencia > 5% — requiere revisión manual"),
        (COLOR_MISSING, "Solo Payway", "El cupón está en Payway pero no en Zetti"),
        (COLOR_MISSING, "Solo Zetti",  "El lote está en Zetti pero no en Payway"),
    ]
    for ri, (bg, estado, desc) in enumerate(leyenda, 2):
        c1 = ws3.cell(row=ri, column=1, value=estado)
        c1.font = Font(name="Arial", bold=True, size=9)
        c1.fill = PatternFill("solid", start_color=bg)
        c1.border = BORDER; c1.alignment = Alignment(horizontal="center")
        c2 = ws3.cell(row=ri, column=2, value=desc)
        c2.font = Font(name="Arial", size=9)
        c2.fill = PatternFill("solid", start_color=bg)
        c2.border = BORDER; c2.alignment = Alignment(horizontal="left")
        ws3.row_dimensions[ri].height = 20

    nota = [
        "",
        "Nota sobre Visa y Mastercard:",
        "Zetti no distingue variantes (Visa Electron, Visa Prepago, Mastercard Debit, etc.).",
        "Este script agrupa todas las variantes Visa bajo 'VISA' y todas las Mastercard bajo 'MASTERCARD'",
        "para permitir la comparación. Diferencias pequeñas pueden deberse a pagos QR desde apps bancarias",
        "(salen como Visa/MC prepago en Payway pero como 'tarjeta modo' en Zetti).",
    ]
    for ri, txt in enumerate(nota, len(leyenda) + 3):
        c = ws3.cell(row=ri, column=1, value=txt)
        c.font = Font(name="Arial", size=8.5,
                      italic=(ri > len(leyenda) + 3),
                      bold=(ri == len(leyenda) + 4))
        ws3.merge_cells(f"A{ri}:B{ri}")

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Punto de entrada del pipeline
# ---------------------------------------------------------------------------
def run(
    payway_path: str,
    zetti_path: str,
    output_path: str,
    fecha: date,
) -> tuple[str, dict]:
    """
    Concilia Payway contra Zetti y genera el Excel final.

    payway_path: ruta al Excel de Payway procesado (paso 2)
    zetti_path:  ruta al CSV de Zetti (paso 3)
    output_path: ruta local donde guardar el Excel de conciliación
    fecha:       date conciliada (para el título del Excel)

    Retorna (output_path, resumen_dict) donde resumen_dict tiene stats para el mail.
    """
    print(f"\n{'='*60}")
    print(f"  PASO 4 — Conciliación — {fecha.strftime('%d/%m/%Y')}")
    print(f"{'='*60}")

    print("  Leyendo Payway (hoja 'Carga de lotes')...")
    df_pay = leer_payway(payway_path)
    print(f"  {len(df_pay)} filas | sucursales: {sorted(df_pay['sucursal'].unique())}")

    print("  Leyendo Zetti...")
    df_zet = leer_zetti(zetti_path)
    print(f"  {len(df_zet)} filas | sucursales: {sorted(df_zet['sucursal'].unique())}")

    print("  Conciliando...")
    df_conc = conciliar(df_pay, df_zet)

    resumen_estados = df_conc["estado"].value_counts().to_dict()
    print(f"  Resultado:")
    for estado, cnt in resumen_estados.items():
        print(f"    {estado:15s}: {cnt}")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    escribir_excel(df_conc, output_path, fecha)
    print(f"  ✅ Excel generado: {os.path.basename(output_path)}")

    # Stats para el mail de éxito
    resumen = {
        "sucursales":        df_conc["sucursal"].nunique(),
        "filas_payway":      int(df_pay["txn_payway"].sum()),
        "cupones_zetti":     int(df_zet["txn_zetti"].sum()),
        "filas_conciliacion": len(df_conc),
        "estados":           resumen_estados,
    }
    return output_path, resumen
