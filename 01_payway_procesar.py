"""
01_payway_procesar.py — Doblen Solutions x Farmacias del Pueblo
1. Lee los 3 CSV de Payway desde la carpeta 1_csvs/
2. Filtra: tipo_operacion=Compra, estado=Aprobado
3. Agrupa tarjetas: VISA + VISA PREPAGO → VISA
                    MASTERCARD + MASTERCARD PREPAGO → MASTERCARD
4. Genera Excel en 2_payway/payway_YYYYMMDD.xlsx con:
   - Hoja "Detalle"        → filas filtradas (Compra + Aprobado)
   - Hoja "Carga de lotes" → pivot sucursal × tarjeta

Correr con: python 01_payway_procesar.py
"""

import os
import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Config de rutas
# ---------------------------------------------------------------------------
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR  = os.path.join(BASE_DIR, "csvs")
OUTPUT_DIR = os.path.join(BASE_DIR, "2_payway")

# ---------------------------------------------------------------------------
# Lookup terminal → (sucursal, integración)
# ---------------------------------------------------------------------------
TERMINALES = {
    16499128: ("central", "desintegrada"), 16499129: ("central", "desintegrada"),
    16499131: ("central", "desintegrada"), 16499132: ("central", "desintegrada"),
    16499619: ("central", "desintegrada"), 16517122: ("central", "desintegrada"),
    16499893: ("central", "desintegrada"), 16499894: ("central", "desintegrada"),
    16499895: ("central", "desintegrada"), 16499896: ("central", "desintegrada"),
    16499897: ("central", "desintegrada"), 16559171: ("central", "desintegrada"),
    16499898: ("belgrano", "desintegrada"), 16489979: ("mitre", "desintegrada"),
    16499899: ("alberdi", "desintegrada"), 16516525: ("j b justo", "desintegrada"),
    16516524: ("j b justo", "desintegrada"), 16500125: ("italia", "desintegrada"),
    16499900: ("italia", "desintegrada"), 16499902: ("plottier", "desintegrada"),
    16499903: ("plottier", "desintegrada"), 16499904: ("plottier", "desintegrada"),
    16499905: ("anonima", "desintegrada"), 16499907: ("perito moreno", "desintegrada"),
    16499908: ("perticone", "desintegrada"), 16499910: ("jumbo", "desintegrada"),
    16499909: ("jumbo", "desintegrada"), 16499911: ("zapala", "desintegrada"),
    16499912: ("zapala", "desintegrada"), 16499913: ("zapala", "desintegrada"),
    16519038: ("cipolletti", "desintegrada"), 16519039: ("cipolletti", "desintegrada"),
    16499915: ("coto", "desintegrada"), 16517123: ("coto", "desintegrada"),
    16527887: ("centenario", "desintegrada"), 16528029: ("centenario", "desintegrada"),
    16499916: ("oeste", "desintegrada"), 16499917: ("oeste", "desintegrada"),
    16519044: ("allen", "desintegrada"), 16499965: ("central", "integrada"),
    16499966: ("central", "integrada"), 16499967: ("central", "integrada"),
    16499968: ("central", "integrada"), 16499969: ("central", "integrada"),
    16507086: ("central", "integrada"), 16584664: ("central", "integrada"),
    16584665: ("central", "integrada"), 16584667: ("central", "integrada"),
    16584668: ("central", "integrada"), 16499958: ("central", "integrada"),
    16499959: ("central", "integrada"), 16499960: ("central", "integrada"),
    16499961: ("central", "integrada"), 16499962: ("central", "integrada"),
    16499973: ("belgrano", "integrada"), 16381950: ("mitre", "integrada"),
    16500000: ("alberdi", "integrada"), 16516550: ("j b justo", "integrada"),
    16516574: ("j b justo", "integrada"), 16500012: ("italia", "integrada"),
    16500011: ("italia", "integrada"), 16500013: ("plottier", "integrada"),
    16500014: ("plottier", "integrada"), 16500016: ("plottier", "integrada"),
    16499979: ("anonima", "integrada"), 16500019: ("perito moreno", "integrada"),
    16534150: ("perticone", "integrada"), 16500089: ("jumbo", "integrada"),
    16500090: ("jumbo", "integrada"), 16500092: ("zapala", "integrada"),
    16500093: ("zapala", "integrada"), 16500094: ("zapala", "integrada"),
    16519116: ("cipolletti", "integrada"), 16519117: ("cipolletti", "integrada"),
    16499993: ("coto", "integrada"), 16528445: ("centenario", "integrada"),
    16438810: ("centenario", "integrada"), 16500098: ("oeste", "integrada"),
    16500099: ("oeste", "integrada"), 16519118: ("allen", "integrada"),
    16559190: ("perticone", "integrada"), 16596482: ("belgrano", "integrada"),
    16596493: ("zapala", "integrada"), 16596494: ("zapala", "integrada"),
    16596495: ("zapala", "integrada"), 16596488: ("anonima", "integrada"),
    16596485: ("j b justo", "integrada"), 16596486: ("j b justo", "integrada"),
    16596490: ("italia", "integrada"), 16596491: ("italia", "integrada"),
    16596483: ("alberdi", "integrada"), 16596484: ("perito moreno", "integrada"),
    16596489: ("mitre", "integrada"), 16605010: ("central", "integrada"),
    16605011: ("central", "integrada"), 16601304: ("plottier", "integrada"),
    16601303: ("plottier", "integrada"), 16601305: ("plottier", "integrada"),
    16608748: ("jumbo", "desintegrada"), 16608749: ("jumbo", "desintegrada"),
    16608750: ("jumbo", "desintegrada"), 16616392: ("j b justo", "desintegrada"),
    16616409: ("coto", "desintegrada"), 16630956: ("cipolletti", "desintegrada"),
    16633767: ("plottier", "desintegrada"), 16633766: ("plottier", "desintegrada"),
    16649451: ("plottier", "desintegrada"), 16649455: ("tarjetas", "desintegrada"),
    16649456: ("tarjetas", "desintegrada"), 16649457: ("tarjetas", "desintegrada"),
    16663334: ("centenario", "desintegrada"), 16657940: ("cipolletti", "desintegrada"),
    16657941: ("cipolletti", "desintegrada"), 16657942: ("cipolletti", "desintegrada"),
    16673067: ("central", "desintegrada"), 16673066: ("central", "desintegrada"),
    16673065: ("central", "desintegrada"), 16689238: ("belgrano", "integrada"),
    16689240: ("central", "integrada"), 16689239: ("central", "integrada"),
    16689241: ("central", "integrada"), 16689242: ("central", "integrada"),
    16689243: ("central", "integrada"), 16689245: ("mitre", "integrada"),
    16689248: ("alberdi", "integrada"), 16689300: ("j b justo", "integrada"),
    16689301: ("j b justo", "integrada"), 16689302: ("italia", "integrada"),
    16689303: ("italia", "integrada"), 16689310: ("anonima", "integrada"),
    16689312: ("perito moreno", "integrada"), 16689315: ("perticone", "integrada"),
    16689317: ("jumbo", "integrada"), 16689318: ("jumbo", "integrada"),
    16689324: ("coto", "integrada"), 16689328: ("oeste", "integrada"),
    16689329: ("oeste", "integrada"), 16690635: ("plottier", "integrada"),
    16690636: ("plottier", "integrada"), 16690634: ("plottier", "integrada"),
    16690641: ("zapala", "integrada"), 16690642: ("zapala", "integrada"),
    16690643: ("zapala", "integrada"), 16690649: ("cipolletti", "integrada"),
    16690650: ("cipolletti", "integrada"), 16690651: ("cipolletti", "integrada"),
    16690652: ("allen", "integrada"), 16711479: ("alto comahue", "desintegrada"),
    16711496: ("alto comahue", "integrada"), 16711884: ("alto comahue", "desintegrada"),
    16711499: ("alto comahue", "integrada"), 16711476: ("alto comahue", "desintegrada"),
    16711497: ("alto comahue", "integrada"), 16680271: ("central", "desintegrada"),
    16711484: ("alto comahue", "desintegrada"), 16726284: ("san martin", "desintegrada"),
    16726356: ("san martin", "integrada"), 16739309: ("anonima", "desintegrada"),
    16772229: ("oeste", "integrada"), 16772233: ("central", "integrada"),
    16772603: ("cipolletti", "integrada"),
}

# ---------------------------------------------------------------------------
# Agrupación de tarjetas
# VISA PREPAGO → VISA
# MASTERCARD PREPAGO → MASTERCARD
# MASTERCARD DEBIT → MASTERCARD DEBITO (normalización)
# VISA DEBITO → VISA DEBITO (se mantiene)
# ---------------------------------------------------------------------------
def agrupar_tarjeta(nombre: str) -> str:
    if pd.isna(nombre):
        return "DESCONOCIDA"
    n = str(nombre).upper().strip()
    # VISA: crédito + prepago → VISA
    if "VISA" in n and "PREPAGO" in n:
        return "VISA"
    if "VISA" in n and not any(x in n for x in ("DEBITO", "DÉBITO", "ELECTRON")):
        return "VISA"
    # VISA DEBITO/ELECTRON se mantiene
    if "VISA" in n:
        return "VISA DEBITO"
    # MASTERCARD: crédito + prepago → MASTERCARD
    if ("MASTERCARD" in n or "MASTER CARD" in n) and "PREPAGO" in n:
        return "MASTERCARD"
    if ("MASTERCARD" in n or "MASTER CARD" in n) and not any(x in n for x in ("DEBIT", "DÉBITO", "DEBITO")):
        return "MASTERCARD"
    # MASTERCARD DEBIT → MASTERCARD DEBITO
    if "MASTERCARD" in n or "MASTER CARD" in n:
        return "MASTERCARD DEBITO"
    return nombre.strip()

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
_thin  = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _hdr(cell, texto=None):
    if texto is not None:
        cell.value = texto
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def _dat(cell, alt=False):
    cell.font      = Font(name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color="F2F2F2" if alt else "FFFFFF")
    cell.border    = BORDER

def _sub(cell):
    cell.font      = Font(name="Arial", bold=True, size=10)
    cell.fill      = PatternFill("solid", start_color="D6E4F0")
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="right")

def _tot(cell):
    cell.font      = Font(name="Arial", bold=True, size=10)
    cell.fill      = PatternFill("solid", start_color="BDD7EE")
    cell.border    = BORDER
    cell.alignment = Alignment(horizontal="right")

# ---------------------------------------------------------------------------
# Lectura y unión de CSVs
# ---------------------------------------------------------------------------
def leer_csvs_local(input_dir: str) -> list:
    if not os.path.isdir(input_dir):
        raise FileNotFoundError(f"No se encontró la carpeta: {input_dir}\nCreála y poné los CSV de Payway adentro.")
    archivos = [f for f in os.listdir(input_dir) if f.lower().endswith(".csv")]
    if not archivos:
        raise FileNotFoundError(f"No hay archivos .csv en {input_dir}")
    print(f"  {len(archivos)} CSV(s) encontrados en 1_csvs/:")
    resultado = []
    for nombre in sorted(archivos):
        path = os.path.join(input_dir, nombre)
        with open(path, "rb") as f:
            data = f.read()
        print(f"    {nombre}")
        resultado.append((nombre, data))
    return resultado


def _lookup_terminal(val):
    try:
        return TERMINALES.get(int(val), ("DESCONOCIDA", "DESCONOCIDA"))
    except (ValueError, TypeError):
        return ("DESCONOCIDA", "DESCONOCIDA")


def leer_y_unir(contenidos: list) -> pd.DataFrame:
    dfs = []
    for nombre, data in contenidos:
        sociedad = os.path.splitext(nombre)[0]
        df = pd.read_csv(io.BytesIO(data), dtype=str, encoding="utf-8-sig")
        df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
        df["sociedad"] = sociedad
        dfs.append(df)
        print(f"    {sociedad}: {len(df)} filas")

    combined = pd.concat(dfs, ignore_index=True)
    combined["fecha_operacion"] = pd.to_datetime(combined["fecha_operacion"], errors="coerce")
    combined["importe"]         = pd.to_numeric(combined["importe"], errors="coerce").fillna(0)
    combined["cuotas"]          = pd.to_numeric(combined["cuotas"],  errors="coerce").fillna(0)

    lookup = combined["nro_terminal"].map(_lookup_terminal)
    combined["sucursal"]    = lookup.map(lambda x: x[0])
    combined["integracion"] = lookup.map(lambda x: x[1])

    sin_lookup = combined[combined["sucursal"] == "DESCONOCIDA"]["nro_terminal"].dropna().unique().tolist()
    if sin_lookup:
        print(f"  ⚠️  Terminales sin lookup: {sin_lookup}")

    # Filtro global: solo Compra + Aprobado
    antes = len(combined)
    combined = combined[
        (combined["tipo_operacion"].str.strip().str.lower() == "compra") &
        (combined["estado"].str.strip().str.lower() == "aprobado")
    ].copy()
    print(f"  Filtro Compra+Aprobado: {antes} → {len(combined)} filas")

    # Agrupación de tarjetas
    combined["marca_tarjeta"] = combined["marca_tarjeta"].apply(agrupar_tarjeta)

    return combined

# ---------------------------------------------------------------------------
# Hoja Detalle
# ---------------------------------------------------------------------------
def escribir_detalle(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title        = "Detalle"
    ws.freeze_panes = "A2"

    orden = ["sociedad","sucursal","integracion","fecha_operacion","hora_operacion",
             "nro_terminal","establecimiento","marca_tarjeta","nro_tarjeta","nombre",
             "tipo_operacion","estado","modo_de_entrada","nro_lote",
             "codigo_autorizacion","nro_cupon","moneda","importe","cuotas"]
    cols = [c for c in orden if c in df.columns]

    etiquetas = {
        "sociedad":"Sociedad","sucursal":"Sucursal","integracion":"Integración",
        "fecha_operacion":"Fecha","hora_operacion":"Hora","nro_terminal":"Terminal",
        "establecimiento":"Establecimiento","marca_tarjeta":"Tarjeta",
        "nro_tarjeta":"Nro Tarjeta","nombre":"Titular","tipo_operacion":"Tipo Op.",
        "estado":"Estado","modo_de_entrada":"Modo","nro_lote":"Lote",
        "codigo_autorizacion":"Autorización","nro_cupon":"Cupón",
        "moneda":"Moneda","importe":"Importe","cuotas":"Cuotas",
    }

    df_out = df[cols].copy()
    df_out["fecha_operacion"] = df_out["fecha_operacion"].dt.strftime("%Y-%m-%d")

    for ci, col in enumerate(cols, 1):
        _hdr(ws.cell(row=1, column=ci), etiquetas.get(col, col))

    for ri, fila in enumerate(df_out.itertuples(index=False), 2):
        alt = ri % 2 == 0
        for ci, val in enumerate(fila, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            _dat(cell, alt)
            if cols[ci-1] == "importe":
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right")

    anchos = {"A":20,"B":14,"C":13,"D":12,"E":10,"F":12,"G":16,"H":20,
              "I":18,"J":22,"K":12,"L":12,"M":14,"N":8,"O":14,"P":10,"Q":8,"R":14,"S":8}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 28

# ---------------------------------------------------------------------------
# Hoja Carga de lotes (pivot sucursal × tarjeta, sin % del Día)
# ---------------------------------------------------------------------------
def escribir_pivot(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Carga de lotes")
    ws.freeze_panes = "A3"

    # El df ya viene filtrado (Compra+Aprobado) y con tarjetas agrupadas
    df_f = df.copy()
    df_f["fecha_str"] = df_f["fecha_operacion"].dt.strftime("%Y-%m-%d")

    pivot = (
        df_f.groupby(["fecha_str", "sucursal", "marca_tarjeta"])
        .agg(transacciones=("importe", "count"), importe_total=("importe", "sum"))
        .reset_index()
        .sort_values(["fecha_str", "sucursal", "marca_tarjeta"])
    )

    # Título (5 columnas, sin % del Día)
    ws.merge_cells("A1:E1")
    t           = ws["A1"]
    t.value     = "Carga de lotes — Compras aprobadas  |  sucursal × tarjeta"
    t.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(["Fecha", "Sucursal", "Tarjeta", "Transacciones", "Importe Total"], 1):
        _hdr(ws.cell(row=2, column=ci), h)
    ws.row_dimensions[2].height = 24

    fila = 3
    for fecha in sorted(pivot["fecha_str"].unique()):
        bloque  = pivot[pivot["fecha_str"] == fecha]
        primera = fila

        for idx, (_, row) in enumerate(bloque.iterrows()):
            alt = idx % 2 == 0
            for ci, val in enumerate([row["fecha_str"], row["sucursal"],
                                      row["marca_tarjeta"], row["transacciones"],
                                      row["importe_total"]], 1):
                cell = ws.cell(row=fila, column=ci, value=val)
                _dat(cell, alt)
                if ci == 5:
                    cell.number_format = "#,##0.00"
                    cell.alignment     = Alignment(horizontal="right")
                elif ci == 4:
                    cell.alignment = Alignment(horizontal="right")
            fila += 1

        # Subtotal del día
        ws.merge_cells(f"A{fila}:C{fila}")
        lbl = ws.cell(row=fila, column=1, value=f"Subtotal {fecha}")
        _sub(lbl); lbl.alignment = Alignment(horizontal="left")

        txn = ws.cell(row=fila, column=4, value=f"=SUM(D{primera}:D{fila-1})")
        _sub(txn); txn.number_format = "#,##0"

        imp = ws.cell(row=fila, column=5, value=f"=SUM(E{primera}:E{fila-1})")
        _sub(imp); imp.number_format = "#,##0.00"

        fila += 1

    # Total general
    ws.merge_cells(f"A{fila}:C{fila}")
    tl = ws.cell(row=fila, column=1, value="TOTAL GENERAL")
    _tot(tl); tl.alignment = Alignment(horizontal="left")

    tt = ws.cell(row=fila, column=4, value=f"=SUM(D3:D{fila-1})")
    _tot(tt); tt.number_format = "#,##0"

    ti = ws.cell(row=fila, column=5, value=f"=SUM(E3:E{fila-1})")
    _tot(ti); ti.number_format = "#,##0.00"

    for col, w in {"A":14, "B":18, "C":22, "D":16, "E":18}.items():
        ws.column_dimensions[col].width = w

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("  Payway — 01 Procesador de CSVs")
    print("  Doblen Solutions x Farmacias del Pueblo")
    print("=" * 60)

    print(f"\nLeyendo CSVs desde 1_csvs/...")
    contenidos = leer_csvs_local(INPUT_DIR)

    print("\nUniendo y procesando...")
    df = leer_y_unir(contenidos)
    print(f"  Total filas (Compra+Aprobado): {len(df)}")

    fechas = sorted(df["fecha_operacion"].dropna().dt.strftime("%Y%m%d").unique())
    print(f"  Fechas: {[f'{f[:4]}-{f[4:6]}-{f[6:]}' for f in fechas]}")
    print(f"  Sucursales: {sorted(df['sucursal'].unique())}")
    print(f"  Tarjetas (agrupadas): {sorted(df['marca_tarjeta'].unique())}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for fecha_slug in fechas:
        fecha_dt = pd.to_datetime(fecha_slug, format="%Y%m%d")
        df_dia   = df[df["fecha_operacion"].dt.date == fecha_dt.date()].copy()

        output_path = os.path.join(OUTPUT_DIR, f"payway_{fecha_slug}.xlsx")
        print(f"\nGenerando 2_payway/payway_{fecha_slug}.xlsx ({len(df_dia)} filas)...")

        wb = Workbook()
        escribir_detalle(wb, df_dia)
        escribir_pivot(wb, df_dia)
        wb.save(output_path)
        print(f"  ✓ Listo")

    print(f"\n{len(fechas)} archivo(s) generado(s) en 2_payway/")
    print("=" * 60)


if __name__ == "__main__":
    main()
